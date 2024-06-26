from typing import Optional

import pandas as pd
import openpyxl


class SpreadsheetManager:
    def __init__(self, data_frames: pd.Series(dtype=object), filename: str) -> None:
        """
        Initializes the SpreadsheetManager object.
        :param data_frames: The data frames to export
        :param filename: The name of the Excel file to create
        """
        self.data_frames = data_frames
        self.filename = filename

    def initialize_spreadsheets(self, hyperlinked_columns: set[str], format_column_widths: Optional[bool] = True) -> None:
        """
        Initializes the Excel spreadsheets.
        :param hyperlinked_columns: The columns to hyperlink
        :param format_column_widths: Whether to format the column widths
        :return:
        """
        self._create_spreadsheets()
        self._apply_hyperlinking(hyperlinked_columns)
        if format_column_widths:
            self._format_excel_columns()

    def _create_spreadsheets(self) -> None:
        """
        Creates the Excel spreadsheets.
        :return:
        """
        with pd.ExcelWriter(f"{self.filename}.xlsx", engine='openpyxl') as writer:
            for key, df in self.data_frames.items():
                # Excel sheet names can't be longer than 31 characters
                sheet_name = key[:31]
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    def _format_excel_columns(self) -> None:
        """
        Formats the Excel columns.
        :return:
        """
        workbook = openpyxl.load_workbook(f"{self.filename}.xlsx")
        for sheet in workbook.worksheets:
            self._format_columns(sheet)
        workbook.save(f"{self.filename}.xlsx")

    def _apply_hyperlinking(self, columns_to_hyperlink: set[str]) -> None:
        """
        Applies hyperlinking to the specified columns.
        :param columns_to_hyperlink: The columns to hyperlink
        :return:
        """
        workbook = openpyxl.load_workbook(f"{self.filename}.xlsx")
        for key, df in self.data_frames.items():
            sheet_name = key[:31]  # Sheet name truncated to 31 characters
            sheet = workbook[sheet_name]
            if not df.empty:
                for column_name in columns_to_hyperlink:
                    if column_name in df.columns:
                        column_index = df.columns.get_loc(column_name) + 1  # Get the Excel column index
                        for row, value in enumerate(df[column_name], start=2):  # Assuming row 1 is header
                            cell = sheet.cell(row=row, column=column_index)
                            cell.value = "LINK"  # Set the display text for the hyperlink
                            cell.hyperlink = value  # Set the actual URL as the hyperlink
                            cell.style = "Hyperlink"  # Optional: Applies the hyperlink style
                    else:
                        print(f"Column {column_name} not found in {key} data frame.")
        workbook.save(f"{self.filename}.xlsx")

    @staticmethod
    def _format_columns(sheet) -> None:
        """
        Formats the columns of the Excel sheet.
        :param sheet: The Excel sheet to format
        :return:
        """
        hyperlinked_columns = {"Item URL", "Photo"}
        hyperlinked_columns_width = 25
        for column_cells in sheet.columns:
            column_title = column_cells[0].value
            if column_title is not None:
                column_width = max(
                    len(str(cell.value)) if column_title not in hyperlinked_columns else hyperlinked_columns_width for
                    cell in column_cells)
                sheet.column_dimensions[column_cells[0].column_letter].width = column_width
